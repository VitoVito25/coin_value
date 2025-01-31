from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table

# Variável global para armazenar a instância do navegador
browser = None

def start_browser():
    """
    Função para iniciar o navegador.
    :return: Instância do navegador.
    """
    global browser  # Utiliza a variável global 'browser'

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    
    service = Service(ChromeDriverManager().install())
    browser = webdriver.Chrome(service=service, options=chrome_options)
    return browser

from selenium.common.exceptions import StaleElementReferenceException

def fetch_coin_data(browser, coin_mapping):
    """
    Função para buscar dados de moedas a partir de um mapeamento de URLs.
    :param browser: Instância do navegador.
    :param coin_mapping: Dicionário com nomes das moedas como chaves e URLs como valores.
    :return: Dicionário com nomes das moedas como chaves e seus valores como valores.
    """
    results = {}
    search_index = 1  # Inicializa o índice de pesquisa em 1
    search_length = len(coin_mapping)  # Corrige o uso de length para len()

    for name, url in coin_mapping.items():
        try:
            print(f"Pesquisando {name}... ({search_index}/{search_length})")
            search_index += 1

            browser.get(url)
            
            # Aguarda até que o valor esteja visível na página (Timeout em 10 segundos)
            wait = WebDriverWait(browser, 30)
            value_element = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="section-coin-overview"]/div[2]/span')))
            
            # Reencontrar o elemento (garante que o elemento não está "obsoleto")
            try:
                coin_value = value_element.text.strip()
            except StaleElementReferenceException:
                value_element = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="section-coin-overview"]/div[2]/span')))
                coin_value = value_element.text.strip()

            # Armazena o resultado no dicionário
            results[name] = coin_value
        
        except TimeoutException as e:
            print(f"Erro ao processar {name}: Tempo de carregamento excedido - {e}")
        except NoSuchElementException as e:
            print(f"Erro ao processar {name}: Elemento não encontrado - {e}")
        except StaleElementReferenceException as e:
            print(f"Erro ao processar {name}: Elemento obsoleto - {e}")
    
    return results

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table

def update_table_in_excel(mapping, file_path="coin_data.xlsx", sheet_name="data", table_name="coins"):
    """
    Atualiza os dados de uma tabela no Excel com base em um dicionário, apagando os dados existentes e inserindo os novos.
    
    :param mapping: Dicionário com nomes das moedas como chaves e valores como valores.
    :param file_path: Caminho do arquivo Excel.
    :param sheet_name: Nome da planilha onde a tabela está localizada.
    :param table_name: Nome da tabela a ser atualizada.
    """
    try:
        # Carrega o arquivo Excel
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"A planilha '{sheet_name}' não existe no arquivo '{file_path}'.")
        
        ws = wb[sheet_name]
        
        # Localiza a tabela pela propriedade ref
        table = None
        for t in ws.tables.values():
            if t.name == table_name:
                table = t
                break
        
        if not table:
            raise ValueError(f"A tabela '{table_name}' não foi encontrada na planilha '{sheet_name}'.")

        # Determina os limites da tabela (intervalo de células)
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        
        # Limpa os dados existentes na tabela (excluindo o cabeçalho)
        for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.value = None
        
        # Insere os novos dados dentro da tabela
        current_row = min_row + 1
        for name, value in mapping.items():
            ws.cell(row=current_row, column=min_col, value=name)  # Nome da moeda
            ws.cell(row=current_row, column=min_col + 1, value=value)  # Valor da moeda
            current_row += 1

        # Atualiza o intervalo da tabela para incluir os novos dados
        table.ref = f"{chr(64 + min_col)}{min_row}:{chr(64 + max_col)}{current_row - 1}"

        # Salva as alterações
        wb.save(file_path)
        print(f"Dados atualizados na tabela '{table_name}' da planilha '{sheet_name}' no arquivo '{file_path}'.")
    except FileNotFoundError:
        print(f"Arquivo '{file_path}' não encontrado.")
    except ValueError as e:
        print(e)
    except Exception as e:
        print(f"Erro inesperado: {e}")


def format_mapping_for_excel(coin_mapping):
    """
    Formata os valores no mapeamento removendo o símbolo de dólar ($),
    excluindo as vírgulas de milhares e substituindo os pontos por vírgulas.

    :param coin_mapping: Dicionário com nomes das moedas como chaves e valores como strings.
                         Exemplo: {"BTC": "$1,234.56", "ETH": "$567.89"}
    :return: Novo dicionário com os valores formatados.
             Exemplo: {"BTC": "1234,56", "ETH": "567,89"}
    """
    formatted_mapping = {}

    print('Formatando dados...')
    
    for name, value in coin_mapping.items():
        if value:
            # Remove o símbolo de dólar
            formatted_value = value.replace("$", "")
            # Remove as vírgulas de milhares
            formatted_value = formatted_value.replace(",", "")
            # Substitui os pontos por vírgulas
            formatted_value = formatted_value.replace(".", ",")
            formatted_mapping[name] = formatted_value
        else:
            formatted_mapping[name] = None  # Caso o valor seja inválido

    return formatted_mapping
