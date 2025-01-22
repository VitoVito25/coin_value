from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook

# Variável global para armazenar a instância do navegador
browser = None

def start_browser():
    """
    Função para iniciar o navegador.
    :return: Instância do navegador.
    """
    global browser  # Utiliza a variável global 'browser'

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    
    service = Service(ChromeDriverManager().install())
    browser = webdriver.Chrome(service=service, options=chrome_options)
    return browser

from selenium.common.exceptions import StaleElementReferenceException

def fetch_coin_data(browser, coin_mapping):
    """
    Função para buscar dados de moedas a partir de um mapeamento de URLs.
    :param browser: Instância do navegador.
    :param coin_mapping: Dicionário com nomes das moedas como chaves e URLs como valores.
    :return: Dicionário com nomes das moedas como chaves e seus valores como valores.
    """
    results = {}
    search_index = 1  # Inicializa o índice de pesquisa em 1
    search_length = len(coin_mapping)  # Corrige o uso de length para len()

    for name, url in coin_mapping.items():
        try:
            print(f"Pesquisando {name}... ({search_index}/{search_length})")
            search_index += 1

            browser.get(url)
            
            # Aguarda até que o valor esteja visível na página (Timeout em 10 segundos)
            wait = WebDriverWait(browser, 10)
            value_element = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="section-coin-overview"]/div[2]/span')))
            
            # Reencontrar o elemento (garante que o elemento não está "obsoleto")
            try:
                coin_value = value_element.text.strip()
            except StaleElementReferenceException:
                value_element = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="section-coin-overview"]/div[2]/span')))
                coin_value = value_element.text.strip()

            # Armazena o resultado no dicionário
            results[name] = coin_value
        
        except TimeoutException as e:
            print(f"Erro ao processar {name}: Tempo de carregamento excedido - {e}")
        except NoSuchElementException as e:
            print(f"Erro ao processar {name}: Elemento não encontrado - {e}")
        except StaleElementReferenceException as e:
            print(f"Erro ao processar {name}: Elemento obsoleto - {e}")
    
    return results


def append_to_excel(mapping, file_path="coin_data.xlsx", sheet_name="data"):
    """
    Adiciona os dados do mapeamento à planilha existente no arquivo Excel.
    :param mapping: Dicionário com nomes das moedas como chaves e valores como valores.
    :param file_path: Caminho do arquivo Excel.
    :param sheet_name: Nome da planilha no arquivo Excel.
    """
    try:
        # Carrega o workbook existente
        wb = load_workbook(file_path)
        
        # Verifica se a planilha existe
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"A planilha '{sheet_name}' não existe no arquivo '{file_path}'.")
        
        ws = wb[sheet_name]  # Seleciona a planilha

        # Insere os dados do mapeamento
        for name, value in mapping.items():
            ws.append([name, value])

        # Salva as alterações no arquivo
        wb.save(file_path)
        print(f"Dados adicionados à planilha '{sheet_name}' no arquivo '{file_path}'.")
    except FileNotFoundError:
        print(f"Arquivo '{file_path}' não encontrado. Certifique-se de que ele existe.")
    except ValueError as e:
        print(e)
